import time
import openpyxl
import pprint

import torch
import numpy as np
import math
import cv2

from yolov6.utils.events import load_yaml
from yolov6.layers.common import DetectBackend
from yolov6.data.data_augment import letterbox
from yolov6.utils.nms import non_max_suppression
class_name =[]
class my_yolov6 :
    def __init__(self,weights,device,yaml,img_size,half):
        self.__dict__.update(locals())

        self.device = device
        self.img_size = img_size
        cuda = self.device != 'cpu' and torch.cuda.is_available()
        self.device = torch.device(f'cuda:{device}' if cuda else 'cpu')
        self.model = DetectBackend(weights, device=self.device)
        self.stride = self.model.stride
        self.class_names = load_yaml(yaml)['names']
        self.img_size = self.check_img_size(self.img_size, s=self.stride)  # check image size
        if half & (self.device.type != 'cpu'):
            self.model.model.half()
            self.half = True
        else:
            self.model.model.float()
            self.half = False

        if self.device.type != 'cpu':
            self.model(torch.zeros(1, 3, *self.img_size).to(self.device).type_as(
                next(self.model.model.parameters())))  # warmup

            # Switch model to deploy status
        self.model_switch(self.model.model, self.img_size)
    @staticmethod
    def plot_box_and_label(image, lw, box, label='', color=(128, 128, 128), txt_color=(255, 255, 255)):
        # Add one xyxy box to image with label
        p1, p2 = (int(box[0]), int(box[1])), (int(box[2]), int(box[3]))
        cv2.rectangle(image, p1, p2, color, thickness=lw, lineType=cv2.LINE_AA)
        if label:
            tf = max(lw - 1, 1)  # font thickness
            w, h = cv2.getTextSize(label, 0, fontScale=lw / 3, thickness=tf)[0]  # text width, height
            outside = p1[1] - h - 3 >= 0  # label fits outside box
            p2 = p1[0] + w, p1[1] - h - 3 if outside else p1[1] + h + 3
            cv2.rectangle(image, p1, p2, color, -1, cv2.LINE_AA)  # filled
            cv2.putText(image, label, (p1[0], p1[1] - 2 if outside else p1[1] + h + 2), 0, lw / 3, txt_color,
                        thickness=tf, lineType=cv2.LINE_AA)

    @staticmethod
    def rescale(ori_shape, boxes, target_shape):
        '''Rescale the output to the original image shape'''
        ratio = min(ori_shape[0] / target_shape[0], ori_shape[1] / target_shape[1])
        padding = (ori_shape[1] - target_shape[1] * ratio) / 2, (ori_shape[0] - target_shape[0] * ratio) / 2

        boxes[:, [0, 2]] -= padding[0]
        boxes[:, [1, 3]] -= padding[1]
        boxes[:, :4] /= ratio

        boxes[:, 0].clamp_(0, target_shape[1])  # x1
        boxes[:, 1].clamp_(0, target_shape[0])  # y1
        boxes[:, 2].clamp_(0, target_shape[1])  # x2
        boxes[:, 3].clamp_(0, target_shape[0])  # y2

        return boxes

    @staticmethod
    def make_divisible(x, divisor):
        # Upward revision the value x to make it evenly divisible by the divisor.
        return math.ceil(x / divisor) * divisor

    def check_img_size(self, img_size, s=32, floor=0):
        """Make sure image size is a multiple of stride s in each dimension, and return a new shape list of image."""
        if isinstance(img_size, int):  # integer i.e. img_size=640
            new_size = max(self.make_divisible(img_size, int(s)), floor)
        elif isinstance(img_size, list):  # list i.e. img_size=[640, 480]
            new_size = [max(self.make_divisible(x, int(s)), floor) for x in img_size]
        else:
            raise Exception(f"Unsupported type of img_size: {type(img_size)}")

        if new_size != img_size:
            print(f'WARNING: --img-size {img_size} must be multiple of max stride {s}, updating to {new_size}')
        return new_size if isinstance(img_size,list) else [new_size]*2

    def model_switch(self, model, img_size):
        ''' Model switch to deploy status '''
        from yolov6.layers.common import RepVGGBlock
        for layer in model.modules():
            if isinstance(layer, RepVGGBlock):
                layer.switch_to_deploy()

    def precess_image(self,img_src, img_size, stride, half):
        '''Process image before image inference.'''
        image = letterbox(img_src, img_size, stride=stride)[0]
        # Convert
        image = image.transpose((2, 0, 1))[::-1]  # HWC to CHW, BGR to RGB
        image = torch.from_numpy(np.ascontiguousarray(image))
        image = image.half() if half else image.float()  # uint8 to fp16/32
        image /= 255  # 0 - 255 to 0.0 - 1.0

        return image, img_src

    def infer(self, source, conf_thres=0.4, iou_thres=0.45, classes=None, agnostic_nms=False, max_det=1000):
        img, img_src = self.precess_image(source, self.img_size, self.stride, self.half)
        img = img.to(self.device)
        # print(img)
        if len(img.shape) == 3:
            img = img[None]
            # expand for batch dim

        pred_results = self.model(img)

        det = non_max_suppression(pred_results, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)[0]

        if len(det):
            det[:, :4] = self.rescale(img.shape[2:], det[:, :4], img_src.shape).round()
            for *xyxy, conf, cls in reversed(det):
                class_num = int(cls)  # integer class
                label = f'{self.class_names[class_num]} {conf:.2f}'
                self.plot_box_and_label(img_src, max(round(sum(img_src.shape) / 2 * 0.003), 2), xyxy, label, color=(255,0,0))
                # print(xyxy)
            img_src = np.asarray(img_src)
        return img_src, det

    def detec_character(self,img_crop):
        img_crop = cv2.resize(img_crop, dsize=None, fx=4, fy=4) #tăng size

        img_character, dett = self.infer(img_crop) # detec cái ảnh được cắt ra, return ra các kí tự mà ảnh có các kí tự

        id = self.class_names # lấy tên của các kí tự
        dett = np.array(dett)
        dett = sorted(dett, key=lambda x: x[1])
        dett = np.array(dett)

        bottom_row = []
        top_row = []
        resuilt = []
        if len(dett)>0 :
            top_row.append(dett[0])
            for i in dett[1:]:
                if i[1] - dett[0][1] > (dett[0][3] - dett[0][1]) * 2 / 3:
                    bottom_row.append(i)
                else:
                    top_row.append(i)

            top_row = sorted(top_row, key=lambda x: x[0])
            bottom_row = sorted(bottom_row, key=lambda x: x[0])

            for x in top_row:
                resuilt.append(id[int(x[5])])
            for x in bottom_row:
                resuilt.append(id[int(x[5])])
        return resuilt
def check_before_save(resuilt) :
    list_not_match =[]
    for i in resuilt:
        if i not in list_not_match:
            list_not_match.append(i)
    return list_not_match

def write_list_to_Excel(sheet, list_2d, start_row, start_col):
    for y, row in enumerate(list_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=list_2d[y][x])
def save_in_excel(path,list_not_match):
    wb = openpyxl.load_workbook(path)
    sheet1 = wb['Sheet1']
    write_list_to_Excel(sheet1,list_not_match,sheet1.max_row+1,1)
    wb.save(path)

get_frame = my_yolov6("best_ckpt.pt", "cpu", "my_train.yaml", 640, False)

character = my_yolov6("best_lcs.pt", "cpu", "character_license_plate.yaml", 640, False)
# img_crop = img_predict[int(det[1][1]):int(det[1][3]),int(det[1][0]):int(det[1][2]),:]

img_crop_list=[]
start_time = time.time()
begin_time = 0
display_time = 2
fps = 0
resuilt_list =[]
# img = cv2.imread('xe4.jpg')
# img =cv2.resize(img,(960,680))
# img_detec ,obj_list= get_frame.infer(img)
# for i in range(len(obj_list)):
#     img_crop = img_detec[int(obj_list[i][1]):int(obj_list[i][3]),int(obj_list[i][0]):int(obj_list[i][2]),:]
#
#     print(1)
#     resuilt = character.detec_character(img_crop)
#
#     if len(resuilt) > 0 and len(resuilt)<10:
#         resuilt_list.append(resuilt)
#         img_crop_list.append(img_crop)
# print(resuilt_list)
# cv2.imshow('aaaa',img_detec)
# cv2.waitKey(0)
# cv2.destroyAllWindows()


cap = cv2.VideoCapture(0)
if not cap.isOpened():
    raise IOError("We cannot open webcam")

while True:
    ret, frame = cap.read()
    # resize our captured frame if we need
    frame = cv2.resize(frame, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_AREA)
    # detect object on our frame

    if(time.time()-begin_time>1):
        frame,obj_list = get_frame.infer(frame)
        print(obj_list)
        if len(obj_list)>0:
            for i in range(len(obj_list)):
                img_crop = frame[int(obj_list[i][1]):int(obj_list[i][3]),int(obj_list[i][0]):int(obj_list[i][2]),:]

                print(1)
                resuilt = character.detec_character(img_crop)

                if len(resuilt) > 7 and len(resuilt)<10:
                    resuilt_list.append(resuilt)
                    img_crop_list.append(img_crop)
                # check_before_save(resuilt)
        begin_time = time.time()
    # show us frame with detection
    cv2.imshow("Web cam input", frame)
    if cv2.waitKey(25) & 0xFF == ord("q"):
        cv2.destroyAllWindows()
        break
    fps += 1
    TIME = time.time() - start_time
    if TIME > display_time:
        print("FPS", fps / TIME)
        fps = 0
        start_time = time.time()
print(resuilt_list)
save_in_excel('license_plate.xlsx',resuilt_list)

cap.release()
cv2.destroyAllWindows()
time.sleep(3)
for img in img_crop_list:
    cv2.imshow('license_plate',img)
    time.sleep(3)
cv2.waitKey(0)
cv2.destroyAllWindows()




# cv2.imshow('adg',img_character)

